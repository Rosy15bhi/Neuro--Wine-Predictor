import openpyxl, unicodedata, re
from openpyxl.styles import Font, PatternFill, Alignment

SRC="/sessions/gracious-optimistic-maxwell/mnt/uploads/catalogo_completo.xlsm"
OUT="/sessions/gracious-optimistic-maxwell/mnt/NEURO PROGETTO/data/catalogo_pulito.xlsx"

WEIGHTS={'eleganza':0.30,'completezza':0.175,'visibilità':0.16,'coerenza':0.145,'design':0.12,'attrattività per giovani':0.10}
# NB: order in file is D..I = eleganza, completezza, visibilità, coerenza, design, attrattività
# but weights were fit per-feature; map by feature name. Reassign correctly:
WEIGHTS={'eleganza':0.10,'completezza':0.175,'visibilità':0.30,'coerenza':0.145,'design':0.12,'attrattività per giovani':0.16}

def strip(s):
    return ''.join(c for c in unicodedata.normalize('NFD',s) if unicodedata.category(c)!='Mn').lower()

BIANCO={'bianco','blanc','chardonnay','greco','fiano','falanghina','vermentino','pecorino','cococciola',
 'malvasia','cortese','garganega','trebbiano','verdicchio','grillo','catarratto','inzolia','zibibbo',
 'sauvignon','riesling','gewurztraminer','traminer','muller','grechetto','passerina','ribolla','friulano',
 'vermentino','vernaccia','carricante','fiano','coda di volpe','asprinio','biancolella','bombino','orvieto',
 'custoza','soave','lugana','gavi','arneis','timorasso','bronner','weiss','blanc'}
ROSSO={'rosso','rouge','nero','aglianico','barolo','barbaresco','nebbiolo','sangiovese','primitivo','negroamaro',
 'montepulciano','cannonau','nerello','frappato','magliocco','gaglioppo','sagrantino','merlot','cabernet',
 'syrah','lambrusco','bonarda','barbera','dolcetto','raboso','teroldego','lagrein','schiava','refosco',
 'aglianico','nero d','nero di','sangiovese','corvina','amarone','valpolicella','chianti','brunello','morellino',
 'bardolino','rosso di','taurasi','salice salentino','copertino','gioia del colle','red'}
ROSATO={'rosato','rose','rosé','cerasuolo','chiaretto'}
SPUMANTE={'spumante','metodo classico','franciacorta','prosecco','champagne','trento doc','talento','brut',
 'extra dry','extra brut','satèn','saten','blanc de blancs','crémant','cremant','asti','lambrusco frizzante'}
DOLCE={'passito','moscato','recioto','vendemmia tardiva','dolce','liquoroso','vin santo','sauternes','marsala','sciacchetra'}

def macro(t):
    if not t: return "Non specificato"
    s=strip(str(t))
    if any(k in s for k in SPUMANTE): return "Spumante"
    if any(k in s for k in DOLCE): return "Dolce/Passito"
    if any(k in s for k in ROSATO): return "Rosato"
    if any(k in s for k in ROSSO): return "Rosso"
    if any(k in s for k in BIANCO): return "Bianco"
    return "Non specificato"

wb=openpyxl.load_workbook(SRC, data_only=True); ws=wb.active
feat=['eleganza','completezza','visibilità','coerenza','design','attrattività per giovani']
fcol={'eleganza':'D','completezza':'E','visibilità':'F','coerenza':'G','design':'H','attrattività per giovani':'I'}

out=openpyxl.Workbook(); o=out.active; o.title="catalogo_pulito"
headers=['brand','nome','tipo','tipo_macro','eleganza','completezza','visibilità','coerenza','design',
 'attrattività_giovani','indice_neuro_calc','indice_neuro_orig','media','giudizio',
 'descrizione_generale','descr_eleganza','descr_completezza','descr_visibilità','descr_coerenza',
 'descr_design','descr_attrattività','ref_foto_fronte','ref_foto_retro']
o.append(headers)

def giudizio(x):
    if x>=9: return "Eccellente"
    if x>=8: return "Ottimo"
    if x>=7: return "Buono"
    if x>=6: return "Sufficiente"
    return "Da migliorare"

macro_count={}
for r in range(2,341):
    vals={f:ws[f"{fcol[f]}{r}"].value for f in feat}
    if not all(isinstance(vals[f],(int,float)) for f in feat):
        idx=None
    else:
        idx=round(sum(vals[f]*WEIGHTS[f] for f in feat),2)
    tm=macro(ws[f"C{r}"].value); macro_count[tm]=macro_count.get(tm,0)+1
    o.append([
        ws[f"A{r}"].value, ws[f"B{r}"].value, ws[f"C{r}"].value, tm,
        vals['eleganza'],vals['completezza'],vals['visibilità'],vals['coerenza'],vals['design'],vals['attrattività per giovani'],
        idx, ws[f"Q{r}"].value, ws[f"R{r}"].value, giudizio(idx) if idx else None,
        ws[f"J{r}"].value, ws[f"K{r}"].value, ws[f"L{r}"].value, ws[f"M{r}"].value, ws[f"N{r}"].value,
        ws[f"O{r}"].value, ws[f"P{r}"].value, ws[f"S{r}"].value, ws[f"T{r}"].value
    ])

# styling header
hdr_fill=PatternFill("solid",fgColor="4A235A"); hdr_font=Font(bold=True,color="FFFFFF")
for c in range(1,len(headers)+1):
    cell=o.cell(1,c); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=Alignment(horizontal='center',vertical='center')
o.freeze_panes="A2"
widths={'A':22,'B':40,'C':30,'D':14,'K':12,'L':12,'M':12,'N':12}
for col,w in {'A':22,'B':42,'C':32,'D':16,'K':16,'L':16,'M':12,'N':14}.items(): o.column_dimensions[col].width=w
out.save(OUT)
print("Saved:", OUT)
print("Macro:", macro_count)
